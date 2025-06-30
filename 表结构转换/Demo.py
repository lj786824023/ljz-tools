import os
import re
import sys
from datetime import datetime

from PySide6 import QtWidgets
from PySide6.QtGui import Qt, QIcon
from PySide6.QtWidgets import QMainWindow, QWidget, QListWidgetItem, QTableWidgetItem, QComboBox, QSizePolicy, QDialog, \
    QTextEdit, QPushButton, QVBoxLayout
from loguru import logger
from openpyxl.reader.excel import load_workbook
from qfluentwidgets import InfoBar, InfoBarPosition, ComboBox, LineEdit, Dialog, TextEdit, MessageBox

from 表结构转换.ConfigFile import ConfigFile
from 表结构转换.TableTransUI import Ui_MainWindow


class MyMainWindow(QMainWindow):
    def __init__(self):
        """需要修改原始tbw_1、tbw_2、tbw_3的类名MyTableWidget.CleverTableWidget()"""
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.setting_file = f"{os.path.dirname(sys.argv[0])}/_internal/aaa_etc/表结构转换规则.xlsx"
        self.workbook = load_workbook(
            f"{os.path.dirname(sys.argv[0])}/_internal/aaa_etc/表结构转换规则.xlsx")  # excel对象
        self.settings = []  # 配置清单
        self.column_types = (
            'bigint', 'blob', 'char', 'clob', 'date', 'datetime', 'decimal', 'float', 'int', 'integer', 'longtext',
            'number', 'nvarchar', 'text', 'time', 'timestamp', 'varchar', 'varchar2')  # 类型列表
        self.column_x_rule = ("长度不变", "长度固定", "长度扩长n倍")  # 长度规则
        self.column_y_rule = ("精度不变", "精度固定", "精度扩长n倍")  # 精度规则

        self.ui.Pivot.currentItemChanged.connect(
            lambda k: self.ui.stackedWidget.setCurrentWidget(self.findChild(QWidget, k)))
        self.ui.btn_add_setting.clicked.connect(self.add_setting)  # 新增配置
        self.ui.btn_del_setting.clicked.connect(self.del_setting)  # 删除配置
        self.ui.lsw_setting.itemClicked.connect(self.show_setting)  # 显示配置
        self.ui.lsw_setting.itemChanged.connect(self.rename_setting)  # 重命名配置
        self.ui.btn_save_setting.clicked.connect(self.save_setting)  # 保存配置

        self.ui.tbw_table.itemChanged.connect(self.split_type)  # item内容变化
        self.ui.btn_trans.clicked.connect(self.translate)  # 开始转换
        self.ui.btn_clear_table.clicked.connect(self.clear_table)  # 开始转换
        self.ui.btn_ddl.clicked.connect(self.show_ddl)  # 开始转换
        self.ui.btn_ex.clicked.connect(self.show_ex)  # 案例数据

        self.init_ui()

    def abc(self, *args, **kwargs):
        print(args)
        print(kwargs)

    def add_setting(self):
        """新增配置"""
        setting_name = '配置'
        db_list = []

        for i in range(1, 100):
            setting_name = f"配置{i}"
            if setting_name not in self.settings:
                break
        self.settings.append(setting_name)  # 变量新增
        # ii = ListWidgetItem("可编辑项")
        item = QListWidgetItem(setting_name)
        item.setFlags(item.flags() | Qt.ItemIsEditable)
        self.ui.lsw_setting.addItem(item)  # 列表新增
        self.ui.cbb_setting.addItem(setting_name)  # 下拉框新增

        sheet = self.workbook['配置清单']
        sheet.cell(row=sheet.max_row + 1, column=1).value = setting_name

        # self.workbook.create_sheet(setting_name)
        # 复制默认配置
        source = self.workbook['默认配置']
        target = self.workbook.copy_worksheet(source)
        target.title = setting_name

        self.workbook.save(self.setting_file)

    def check_enable(self):
        """检查工具是否到期"""
        current_date = datetime.now().date()  # 获取当前日期
        enable_date = datetime.strptime("2025-12-31", "%Y-%m-%d").date()  # 工具有效期
        if current_date > enable_date:
            # QMessageBox.information(None, "消息", "已到使用有效期！", QMessageBox.Close)
            w = MessageBox("消息", "已到使用有效期！", self)
            if w.exec():
                self.close()
                sys.exit()
            else:
                self.close()
                sys.exit()

    def clear_table(self):
        """清空页面"""
        for row in range(self.ui.tbw_table.rowCount()):
            for column in [i for i in range(self.ui.tbw_table.columnCount()) if i not in [5, 6, 7]]:
                self.ui.tbw_table.setItem(row, column, QTableWidgetItem(''))

    def createErrorInfoBar(self, text):
        w = InfoBar.error(
            title='ERROR',
            content=text,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.BOTTOM,
            duration=2000,
            parent=self
        )
        # w.addWidget(PushButton('Action'))
        w.show()

    def createInfoInfoBar(self, text):
        w = InfoBar.info(
            title='INFO',
            content=text,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.BOTTOM,
            duration=2000,
            parent=self
        )
        # w.addWidget(PushButton('Action'))
        w.show()

    def createSuccessInfoBar(self, text):
        InfoBar.success(
            title='SUCCESS',
            content=text,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.BOTTOM,
            # position='Custom',   # NOTE: use custom info bar manager
            duration=2000,
            parent=self
        )

    def del_setting(self):
        """删除配置"""
        item = self.ui.lsw_setting.currentItem()
        if not item:
            return

        if item.text() == '默认配置':
            self.createInfoInfoBar("默认配置不可删除！")
            return

        row = self.ui.lsw_setting.currentRow()
        text = item.text()
        self.ui.lsw_setting.takeItem(row)  # 从列表框中移除
        for i in range(self.ui.cbb_setting.count()):  # 从下拉框中移除
            if self.ui.cbb_setting.itemText(i) == text:
                self.ui.cbb_setting.removeItem(i)
                break
        self.settings.remove(text)  # 从变量中移除

        sheet = self.workbook["配置清单"]
        # sheet.delete_rows(1,sheet.max_row)
        for i in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=i, column=1)
            if cell.value == text:
                sheet.delete_rows(cell.row, 1)  # 从某行开始删1行
                break
        del self.workbook[text]  # 删除工作表
        self.workbook.save(self.setting_file)

    def init_ui(self):
        """初始化ui"""
        window_icon = QIcon(f"{os.path.dirname(sys.argv[0])}/_internal/aaa_etc/weixinshoucang.ico")
        # 设置窗口图标
        self.setWindowIcon(window_icon)
        # 读取配置清单
        # self.wb = load_workbook(self.setting_file)
        st = self.workbook["配置清单"]
        self.settings = [row[0] for row in st.iter_rows(min_row=1, max_col=1, values_only=True)]
        self.settings.sort(key=lambda x:x.lower()) # 不区分大小写排升序

        # 初始化配置清单
        self.ui.lsw_setting.clear()
        for setting in self.settings:
            item = QListWidgetItem(setting)
            item.setFlags(item.flags() | Qt.ItemIsEditable) # 可编辑
            self.ui.lsw_setting.addItem(item)

        # 初始化下拉框
        self.ui.cbb_setting.clear()
        self.ui.cbb_setting.addItems(self.settings)
        # for setting in self.settings:
        #     self.ui.cbb_setting.addItem(setting)

        # self.ui.tbw_table.tbw_table.setRowCount(100)
        # self.ui.tbw_table.tbw_table.setColumnCount(9)
        # self.ui.tbw_table.tbw_table.setHorizontalHeaderLabels(('表英文名','表中文名','字段英文名','字段中文名','字段类型','长度','精度','转换后长度','转换后精度'))

        self.ui.tbw_setting.setRowCount(len(self.column_types))  # 设置配置表格行数

    def rename_setting(self, item: QListWidgetItem):
        """重命名配置"""
        text = item.text()
        row = self.ui.lsw_setting.currentRow()

        self.ui.cbb_setting.setItemText(row, text)  # 重命名下拉框数据

        sheet = self.workbook['配置清单']
        sheet.cell(row=row + 1, column=1).value = text  # 重命名配置清单

        self.workbook[self.settings[row]].title = text  # 重命名配置明细
        self.settings[row] = text  # 重命名变量

        self.workbook.save(self.setting_file)

    def save_setting(self):
        """保存配置"""
        item = self.ui.lsw_setting.currentItem()
        if not item:
            self.createInfoInfoBar("未选中配置！")
            return

        sheet = self.workbook[item.text()]
        sheet.delete_rows(2, sheet.max_row)  # 清空

        for row in range(self.ui.tbw_setting.rowCount()):
            text1 = self.ui.tbw_setting.item(row, 0).text()
            text2 = self.ui.tbw_setting.cellWidget(row, 1).currentText()
            text3 = self.ui.tbw_setting.cellWidget(row, 2).currentText()
            text4 = self.ui.tbw_setting.item(row, 3).text()
            text5 = self.ui.tbw_setting.cellWidget(row, 4).currentText()
            text6 = self.ui.tbw_setting.item(row, 5).text()
            text7 = self.ui.tbw_setting.item(row, 6).text()  # 默认长度
            text8 = self.ui.tbw_setting.item(row, 7).text()  # 默认精度
            line = (text1, text2, text3, text4, text5, text6, text7, text8)
            # logger.info(line)
            sheet.append(line)

        self.workbook.save(self.setting_file)

        try:
            self.workbook.save(self.setting_file)
            self.createSuccessInfoBar('保存成功！')
        except:
            self.createErrorInfoBar('保存失败！文件可能已经打开')

    def show_ddl(self):
        """弹出窗口展示ddl"""
        # 创建对话框

        text_tab = self.ui.tbw_table.item(0, 0).text() if self.ui.tbw_table.item(0, 0) else ''  # 表英文名
        text_tab_comment = self.ui.tbw_table.item(0, 1).text() if self.ui.tbw_table.item(0, 0) else ''  # 表中文名
        ddl_text = f"create table {text_tab}("
        for row in range(self.ui.tbw_table.rowCount()):

            text_col = self.ui.tbw_table.item(row, 5).text() if self.ui.tbw_table.item(row, 5) else ''
            if not text_col:  # 末尾结束
                break

            text_col = self.ui.tbw_table.item(row, 2).text() if self.ui.tbw_table.item(row, 2) else ''
            text_col_type = self.ui.tbw_table.item(row, 11).text() if self.ui.tbw_table.item(row, 11) else ''
            text_col_comment = self.ui.tbw_table.item(row, 3).text() if self.ui.tbw_table.item(row, 3) else ''
            line = f"\n  {text_col} {text_col_type} comment '{text_col_comment}',"
            ddl_text += line

        ddl_text = ddl_text[:-1]  # 去掉字段最后的逗号
        ddl_text += f"\n) comment '{text_tab_comment}' ;"

        dialog = QDialog(self)
        dialog.setWindowTitle("生成ddl")
        dialog.setGeometry(150, 150, 500, 400)

        # 创建多行编辑框
        text_edit = TextEdit()
        # text_edit.setPlaceholderText("在这里输入多行文本...")
        text_edit.setText(ddl_text)

        # 创建关闭按钮
        # close_button = QPushButton("关闭")
        # close_button.clicked.connect(dialog.close)

        # 设置布局
        layout = QVBoxLayout()
        layout.addWidget(text_edit)
        # layout.addWidget(close_button)
        dialog.setLayout(layout)

        # 显示对话框
        dialog.exec()

    def show_ex(self):
        """案例数据"""
        tab = 'datamapping'
        tab_comment = '映射表'
        col = ['seq_num', 't_tab_eng_name', 't_tab_chn_name', 't_col_eng_name', 't_col_chn_name', 't_col_datatype',
               't_col_desc', 't_his_flag', 't_desc', 's_group_id', 's_system', 's_tab_eng_name', 's_tab_chn_name',
               's_col_eng_name', 's_col_chn_name', 's_col_datatype', 's_tab_pk', 's_col_desc', 's_col_code_desc',
               'e_priority', 'e_trans_rule', 'e_desc', 'r_trans_exp', 'r_val_rule', 'r_is_join',
               'r_join_source_tabname', 'r_join_market_tabname', 'r_join_type', 'r_join_condition', 'r_where_condition',
               'r_exec_vt', 'r_desc', 'updatetime']
        col_comment = ['序号', '表英文名', '表中文名', '字段英文名', '字段中文名', '字段类型', '字段描述', '历史标志',
                       '表描述', '分组编号', '系统编号', '源表英文名', '源表英文名', '源字段英文名', '源字段中文名',
                       '源字段类型', '源字段主键', '源字段描述', '源字段代码描述', '有效标志', '转换规则', '转换描述',
                       '转换表达式', '字段规则', '是否关联', '关联源系统表名', '关联集市表名', '关联类型', '关联条件',
                       'where条件', '额外执行sql', '备注', '更新时间']
        col_type = ['int', 'varchar2(56)', 'varchar2(100)', 'varchar2(64)', 'varchar2(100)', 'varchar2(60)', 'text',
                    'varchar2(10)', 'text', 'number(12,2)', 'varchar2(64)', 'varchar2(56)', 'varchar2(100)',
                    'varchar2(64)', 'varchar2(1000)', 'varchar2(60)', 'varchar2(1)', 'text', 'text', 'varchar2(10)',
                    'text', 'text', 'text', 'varchar2(500)', 'char(1)', 'text', 'text', 'varchar2(256)', 'text', 'text',
                    'text', 'text', 'timestamp']

        self.clear_table()  # 清空
        for i in range(len(col)):
            self.ui.tbw_table.setItem(i, 0, QTableWidgetItem(tab))  # 表英文名
            self.ui.tbw_table.setItem(i, 1, QTableWidgetItem(tab_comment))  # 表中文名
            self.ui.tbw_table.setItem(i, 2, QTableWidgetItem(col[i]))  # 字段英文名
            self.ui.tbw_table.setItem(i, 3, QTableWidgetItem(col_comment[i]))  # 字段中文名
            self.ui.tbw_table.setItem(i, 4, QTableWidgetItem(col_type[i]))  # 字段类型

    def show_setting(self, item: QListWidgetItem):
        """
        配置点击
        查找所有sheet，如果有直接读取数据，如果没有创建新的sheet
        """
        setting = item.text()
        sheet = self.workbook[setting]
        excel_data = list(sheet.values)[1:]
        self.ui.tbw_setting.clearContents()
        for row in range(len(excel_data)):
            data = QTableWidgetItem(str(excel_data[row][0] or ""))  # 第1列 源类型
            data.setFlags(Qt.ItemIsEnabled)  # 禁用编辑标志
            self.ui.tbw_setting.setItem(row, 0, data)

            combo = ComboBox()  # 第2列 目标类型
            combo.addItems(self.column_types)
            combo.setCurrentIndex(self.column_types.index(excel_data[row][1]))
            self.ui.tbw_setting.setCellWidget(row, 1, combo)

            combo = ComboBox()  # 第3列 长度转换规则
            combo.addItems(self.column_x_rule)
            combo.setCurrentIndex(self.column_x_rule.index(excel_data[row][2]))
            self.ui.tbw_setting.setCellWidget(row, 2, combo)

            item = QTableWidgetItem(str(excel_data[row][3] or ""))  # # 第4列 长度转换值
            self.ui.tbw_setting.setItem(row, 3, item)

            combo = ComboBox()  # 第5列 精度转换规则
            combo.addItems(self.column_y_rule)
            combo.setCurrentIndex(self.column_y_rule.index(excel_data[row][4]))
            self.ui.tbw_setting.setCellWidget(row, 4, combo)

            item = QTableWidgetItem(str(excel_data[row][5] or ""))  # 第6列 精度转换值
            self.ui.tbw_setting.setItem(row, 5, item)

            item = QTableWidgetItem(str(excel_data[row][6] or ""))  # 第7列 长度度默认值
            self.ui.tbw_setting.setItem(row, 6, item)

            item = QTableWidgetItem(str(excel_data[row][7] or ""))  # 第8列 精度默认值
            self.ui.tbw_setting.setItem(row, 7, item)

    def split_type(self, item: QTableWidgetItem):
        """拆解类型"""
        row = item.row()
        column = item.column()
        text = item.text()

        if column != 4:  # 只有当字段类型发生变化时才执行
            return

        new_text = re.split(r"\(|,|\)", text) + ['', '', '']  # 按照(,)拆分
        self.ui.tbw_table.setItem(row, column + 1, QTableWidgetItem(new_text[0]))  # 拆解类型
        self.ui.tbw_table.setItem(row, column + 2, QTableWidgetItem(new_text[1]))  # 拆解长度
        self.ui.tbw_table.setItem(row, column + 3, QTableWidgetItem(new_text[2]))  # 拆解精度

    def translate(self):
        """开始转换"""
        # 获取当前配置信息
        setting = self.ui.cbb_setting.currentText()
        sheet = self.workbook[setting]

        setting_detail = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]  # 获取配置明细

        for row in range(self.ui.tbw_table.rowCount()):
            item_type = self.ui.tbw_table.item(row, 5)
            if not item_type:  # 没有值了跳出
                break

            # 清空后3列
            # for row in range(self.ui.tbw_table.rowCount()):
            #     self.ui.tbw_table.setItem(row,8,QTableWidgetItem(''))
            #     self.ui.tbw_table.setItem(row,9,QTableWidgetItem(''))
            #     self.ui.tbw_table.setItem(row,10,QTableWidgetItem(''))

            item_x = self.ui.tbw_table.item(row, 6)  # 长度
            item_y = self.ui.tbw_table.item(row, 7)  # 精度
            text = item_type.text()  # 字段类型
            for line in setting_detail:
                if line[0].lower() == text.lower():

                    col_type = line[1]
                    self.ui.tbw_table.setItem(row, 8, QTableWidgetItem(col_type))  # 转换后类型

                    if line[2] == '长度不变':
                        x = item_x.text()
                    if line[2] == '长度固定':
                        x = line[3] or ''
                    if line[2] == '长度扩长n倍':
                        x = str(int(float(item_x.text()) * float(line[3] or '1')))
                    self.ui.tbw_table.setItem(row, 9, QTableWidgetItem(x))  # 转换后长度

                    if line[4] == '精度不变':
                        y = item_y.text()
                    if line[4] == '精度固定':
                        y = line[5] or ''
                    if line[4] == '精度扩长n倍':
                        y = str(int(float(item_y.text()) * float(line[5] or '1')))
                    self.ui.tbw_table.setItem(row, 10, QTableWidgetItem(y))  # 转换后精度

                    result = '6'

                    if not x and not y:  # 长度精度同时为空
                        result = col_type  # decimal
                        """
                        if text.lower() == 'number':
                            x_default = line[6]  # 默认长度
                            y_default = line[7]  # 默认精度
                            if not x_default:  # 如果默认长度为空
                                result = col_type  # decimal
                            else:
                                if not y_default:  # 如果默认精度为空
                                    result = f"{col_type}({x_default})"  # decimal(n)
                                else:
                                    result = f"{col_type}({x_default},{y_default})"  # decimal(n,m)
                        """


                    if x and not y:  # 精度为空
                        result = f"{col_type}({x})"
                    if x and y:
                        result = f"{col_type}({x},{y})"
                    if not x and y:
                        result = 'error'
                    self.ui.tbw_table.setItem(row, 11, QTableWidgetItem(result))

                    # break

                    # number类型长度、精度为空特殊处理
                    if text.lower() == 'number' and self.ui.tbw_table.item(row, 6).text() =='' and self.ui.tbw_table.item(row, 7).text() == '': # number类型 原始长度精度为空

                        self.ui.tbw_table.setItem(row, 9, QTableWidgetItem(line[6] or ""))  # 设置转换后长度
                        self.ui.tbw_table.setItem(row, 10, QTableWidgetItem(line[7] or ""))  # 设置转换后精度

                        if line[6] : # 如果默认长度有值
                            if line[7] : # 如果默认精度有值
                                result = f"{line[1]}({line[6]},{line[7]})" # decimal(12,2)
                            else:
                                result = f"{line[1]}({line[6]})"  # decimal(12)
                            # self.ui.tbw_table.setItem(row, 11, QTableWidgetItem(result))  # 转换后精度
                        else:
                            result = col_type # decimal

                        self.ui.tbw_table.setItem(row, 11, QTableWidgetItem(result))  # 设置最终结果




if __name__ == "__main__":
    # pyinstaller.exe .\Demo.py -w --paths C:\Users\lojn\PycharmProjects\ljz-tools --add-data .\_internal\aaa_etc\*:aaa_etc -i C:\Users\lojn\PycharmProjects\DataView\img\weixinshoucang.ico
    # -w 隐藏cmd窗口执行
    # --paths 额外指定import的搜索路径
    # --paths 把指定文件放入打包后的目标文件夹内
    app = QtWidgets.QApplication(sys.argv)
    ui = MyMainWindow()
    ui.show()
    ui.check_enable() # 检查有效期
    sys.exit(app.exec())
