import os
from openpyxl import load_workbook
from openpyxl.styles import Font, colors

font = Font(color=colors.BLUE, underline="single")


def create_dir_hyperlink(file_dir, return_index):
    """生成目录超链接"""
    workbook = load_workbook(file_dir)
    sheet_names = workbook.sheetnames
    if "目录索引" in sheet_names:
        workbook.remove(workbook["目录索引"])
    workbook.create_sheet("目录索引", 0)
    sheet = workbook["目录索引"]

    sheet_names = workbook.sheetnames
    for i in range(1, len(sheet_names)):
        sheet[f"A{str(i)}"] = sheet_names[i]  # sheet名
        # =HYPERLINK("#A!A1","a")
        sheet[f"B{str(i)}"] = f'=HYPERLINK("#{sheet_names[i]}!A1","跳转")'  # 添加跳转超链接
        sheet[f"B{str(i)}"].font = font  # 设置文字样式

        workbook[sheet_names[i]][return_index] = f'=HYPERLINK("#目录索引!B{str(i)}","返回")'  # 添加返回超链接
        workbook[sheet_names[i]][return_index].font = font  # 设置文字样式

    workbook.save(file_dir) # 可能会有异常（文件已打开）


if __name__ == "__main__":
    comment = """
********************************************************************************
功能：自动为每个sheet添加跳转、返回超链接
  新建一个“目录索引”的sheet，A列为所有sheet名，B列为跳转到该sheet的超链接
  为每个sheet添加一个返回到“目录索引”的超链接
********************************************************************************
"""
    print(comment)
    file_dir = input(
        "1.请输入Excel完整路径：")  # C:\Users\lojn\PycharmProjects\ljz-tools\工具脚本\Excel生成跳转返回超链接\dist\test.xlsx
    file_dir=file_dir.replace('"',"")
    # 检查文件是否存在
    while not os.path.exists(file_dir):
        # print(file_dir)
        file_dir = input("文件不存在，请重新输入：")

    return_index = input("2.请输入生成返回超链接的列（例：A1、B1、C1，不填为G1）：")
    return_index = "G1" if not return_index else return_index

    # "C:\\Users\\lojn\\PycharmProjects\\ljz-tools\\工具脚本\\Excel生成跳转返回超链接\\dist\\test.xlsx"
    create_dir_hyperlink(file_dir, return_index)

    while True:
        res = input("3.已生成超链接，输入回车退出。")
        if not res:
            break
