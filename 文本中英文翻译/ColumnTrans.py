import os
import sys

import openpyxl
from PySide6 import QtWidgets
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QMainWindow, QTableWidgetItem
from openpyxl.reader.excel import load_workbook
from ui.ColTransUI import Ui_MainWindow


class MyMainWindow(QMainWindow):
    def __init__(self):
        """需要修改原始tbw_1、tbw_2、tbw_3的类名MyTableWidget.CleverTableWidget()"""
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.btn_trans.clicked.connect(self.trans_columns)  # 按钮绑定
        self.ui.btn_save.clicked.connect(self.save_file)  # 按钮绑定
        self.dir_file = os.path.dirname(sys.argv[0]) + "/_internal/aaa_etc/column_trans.xlsx"  # 配置文件路径：当前
        self.ui.label.setText(self.dir_file)
        self.load_file()  # 加载文件内容
        # 表头居左
        self.ui.tbw_1.horizontalHeader().setDefaultAlignment(Qt.AlignLeft)
        self.ui.tbw_2.horizontalHeader().setDefaultAlignment(Qt.AlignLeft)
        self.ui.tbw_3.horizontalHeader().setDefaultAlignment(Qt.AlignLeft)

    def save_file(self):
        """保存文件"""
        wb = openpyxl.Workbook()  # 新建工作簿
        sheet = wb.create_sheet("字段列表", 0)  # 新建sheet
        sheet["A1"] = "字段中文名"
        sheet["B1"] = "生成中文名"
        sheet["C1"] = "生成英文名"
        sheet["D1"] = "调整后"
        sheet = wb.create_sheet("修改记录", 1)  # 新建sheet
        sheet["A1"] = "修改日期"
        sheet["B1"] = "修改人"
        sheet["C1"] = "修改说明"
        sheet["D1"] = "修改原因"
        sheet["E1"] = "备注"
        sheet = wb.create_sheet("术语汇总", 2)  # 新建sheet
        sheet["A1"] = "名词中文名"
        sheet["B1"] = "名称英文名"
        sheet["C1"] = "使用次数"
        for index, tbw in enumerate([self.ui.tbw_1, self.ui.tbw_2, self.ui.tbw_3]):  # 遍历3个表单
            sheet_key = self.ui.tabWidget.tabText(index)  # 字段列表、修改记录、术语汇总
            sheet = wb[sheet_key]  # 获取sheet
            for row in range(tbw.rowCount()):  # 遍历行
                for col in range(tbw.columnCount()):  # 遍历列
                    item = tbw.item(row, col)
                    if item:
                        text = item.text()
                    else:
                        text = ""
                    sheet.cell(row + 2, col + 1, text)  # 写入
        sheet["A1"] = "text"
        wb.save(self.dir_file)  # 执行保存

    def load_file(self):
        wb = load_workbook(self.dir_file)
        # 读取字段列表
        sheet = wb["字段列表"]
        max_row = sheet.max_row
        self.ui.tbw_1.setRowCount(max_row - 1)  # 忽略首行
        self.ui.tbw_1.setRowCount(max_row - 1)  # 忽略首行
        self.data_1 = [line for line in sheet.rows]
        for index_row, line in enumerate(self.data_1[1:]):
            for index_col, cell in enumerate(line):
                item = QTableWidgetItem(str(cell.value or ""))
                self.ui.tbw_1.setItem(index_row, index_col, item)
        self.ui.tbw_1.resizeColumnsToContents()  # 单元格自适应

        # 读取修改记录
        sheet = wb["修改记录"]
        max_row = sheet.max_row
        self.ui.tbw_2.setRowCount(max_row - 1)
        self.ui.tbw_2.setRowCount(10)
        self.data_2 = [line for line in sheet.rows]
        for index_row, line in enumerate(self.data_2[1:]):
            for index_col, cell in enumerate(line):
                item = QTableWidgetItem(str(cell.value or ""))
                self.ui.tbw_2.setItem(index_row, index_col, item)
        self.ui.tbw_2.resizeColumnsToContents()  # 单元格自适应

        # 读取术语汇总
        sheet = wb["术语汇总"]
        max_row = sheet.max_row
        self.ui.tbw_3.setRowCount(max_row - 1)
        self.data_3 = []
        self.trans_dicts = {}
        for row in sheet.rows:
            line = [str(x.value or "") for x in row]
            self.data_3.append(line)
        del self.data_3[0]
        for row in range(len(self.data_3)):
            for col in range(len(self.data_3[row])):
                item = QTableWidgetItem(str(self.data_3[row][col]))
                self.ui.tbw_3.setItem(row, col, item)
            self.trans_dicts[self.data_3[row][0]] = self.data_3[row][1]
        self.ui.tbw_3.resizeColumnsToContents()  # 单元格自适应

    def trans_columns(self):
        # 获取字段列表翻译清单
        for row in range(self.ui.tbw_1.rowCount()):
            item = self.ui.tbw_1.item(row, 0)
            if item:
                trans_result = self.trans_text([str(item.text()), "", ""])
                item_1 = QTableWidgetItem(trans_result[2])
                item_2 = QTableWidgetItem(trans_result[1])
                self.ui.tbw_1.setItem(row, 1, item_1)
                self.ui.tbw_1.setItem(row, 2, item_2)

            else:  # 如果为空则跳出循环
                break
        self.ui.tbw_1.resizeColumnsToContents()

    def trans_text(self, text: list = []):
        if text[0] == "":
            return text
        str_len = len(text[0])
        # 找到
        for i in range(str_len, 0, -1):
            find_text = text[0][:i]
            if find_text in self.trans_dicts:
                if str_len == i:
                    text[0] = ""
                    text[1] = text[1] + self.trans_dicts[find_text]
                    text[2] = text[2] + find_text
                else:
                    text[0] = text[0][i:]
                    text[1] = text[1] + self.trans_dicts[find_text] + "_"
                    text[2] = text[2] + find_text + "_"
                return self.trans_text(text)
        # 一个都没找到
        if str_len == 1:
            text[1] = text[1] + "??"
            text[2] = text[2] + "?" + text[0][0:1] + "?"
            text[0] = ""
        else:
            text[0] = text[0][1:]
            text[1] = text[1] + "??_"
            text[2] = text[2] + "?" + find_text + "?_"
        return self.trans_text(text)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    ui = MyMainWindow()
    ui.show()
    sys.exit(app.exec())
