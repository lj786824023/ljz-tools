import os
import sys
import time

import pymysql
from dbutils.pooled_db import PooledDB
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, colors
from openpyxl.workbook import Workbook


def gbase_pool():
    return PooledDB(
        creator=pymysql,  # 使用链接数据库的模块
        maxconnections=6,  # 连接池允许的最大连接数
        mincached=2,  # 初始化时，链接池中至少创建的空闲的链接
        maxcached=5,  # 链接池中最多闲置的链接
        maxshared=3,  # 链接池中最多共享的链接数量
        blocking=True,  # 连接池中如果没有可用连接后，是否阻塞等待
        host='172.16.30.34',  # 主机号
        port=3306,  # 端口号
        user='query',  # 用户名
        password='query',  # 密码
        database='etl',  # schema
        charset='utf8',  # 数据库编码
        cursorclass=pymysql.cursors.SSCursor
    )
    # return PooledDB(
    #     creator=pymysql,  # 使用链接数据库的模块
    #     maxconnections=6,  # 连接池允许的最大连接数
    #     mincached=2,  # 初始化时，链接池中至少创建的空闲的链接
    #     maxcached=5,  # 链接池中最多闲置的链接
    #     maxshared=3,  # 链接池中最多共享的链接数量
    #     blocking=True,  # 连接池中如果没有可用连接后，是否阻塞等待
    #     host='172.16.30.34',  # 主机号
    #     port=3306,  # 端口号
    #     user='query',  # 用户名
    #     password='query',  # 密码
    #     database='etl',  # schema
    #     charset='utf8',  # 数据库编码
    #     cursorclass=pymysql.cursors.SSCursor
    # )


if __name__ == "__main__":

    # 查询datamapping_task
    pool = gbase_pool()
    db = pool.connection()
    cursor = db.cursor()

    c1 = db.cursor()
    c2 = db.cursor()
    c1.execute("select count(1) as rn from datamapping_task limit 5")
    c2.execute("select count(1) as rn from datamapping_task limit 10")
    s1=c1.fetchall()
    s2=c2.fetchall()
    print(s1)
    print(s2)
    print(666)

    cursor.execute("select * from datamapping_task order by t_tab_eng_name")
    datamapping_task_head = [line[0] for line in cursor.description] if cursor.description else []
    datamapping_task = cursor.fetchall()
    print(f"datamapping数据量：{len(datamapping_task)}")
    cursor.execute("select * from datamapping order by 2,1")
    datamapping_head = [line[0] for line in cursor.description] if cursor.description else []
    datamapping = cursor.fetchall()
    print(f"datamapping数据量：{len(datamapping)}")
    # print(datamapping)

    # 生成excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "目录"
    # sheet = workbook.create_sheet("目录")

    # 写入datamapping_task
    sheet.append(datamapping_task_head)
    for i, line in enumerate(datamapping_task):
        sheet.append(line)
        sheet[f"A{i + 2}"] = f'=HYPERLINK("#{line[1]}!A1","跳转")'  # 添加跳转超链接
        sheet[f"A{i + 2}"].font = Font(color=colors.BLUE, underline="single")  # 设置文字样式

        # cell = WriteOnlyCell(sheet,value=f'=HYPERLINK("#{line[1]}!A1","跳转")')
        # cell.font = Font(color=colors.BLUE, underline="single")  # 设置文字样式
        # print(type([cell]),type(list(line[1:])))
        # c=[cell]+list(line[1:])
        # sheet.append([cell]+list(line)[1:])

    f = Font(color=colors.BLUE, underline="single")  # 设置文字样式
    # 写入datamapping
    for i, line in enumerate(datamapping_task):
        sheet = workbook.create_sheet(line[1])
        sheet.append(datamapping_head)
        sheet[f"A1"] = f'=HYPERLINK("#目录!B{i + 2}","返回")'  # 添加跳转超链接
        sheet[f"A1"].font = Font(color=colors.BLUE, underline="single")  # 设置文字样式
        # cell = WriteOnlyCell(sheet, value=f'=HYPERLINK("#目录!B{i + 2}","返回")')
        # cell.font = f
        # sheet.append([cell] + list(datamapping_head)[1:])

        for data in datamapping:
            if line[1] == data[1]:
                sheet.append(data)
        print(f"{line[1]}已完成")

    print("正在写入文件...")
    workbook.save(f"映射文档.xlsx")
    # os.path.dirname(sys.argv[0]) + "/_internal/aaa_etc/enviro
    print(os.path.dirname(sys.argv[0]) + "\映射文档.xlsx")
    while True:
        res = input("输入回车退出:")
        if not res:
            break

